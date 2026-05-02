# ============================================================
#           WebEarl Technologies — Excel Writer
# ============================================================

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from config import OUTPUTS_FOLDER, EXCEL_FILE_NAME, EXCEL_COLUMNS


# --- Excel File Path ---
EXCEL_PATH = os.path.join(OUTPUTS_FOLDER, EXCEL_FILE_NAME)

# --- Colors ---
HEADER_BG_COLOR = "1F4E79"       # Dark blue
HEADER_FONT_COLOR = "FFFFFF"     # White
ROW_ODD_COLOR = "DCE6F1"         # Light blue
ROW_EVEN_COLOR = "FFFFFF"        # White
INTERNSHIP_COLOR = "E2EFDA"      # Light green
FULLTIME_COLOR = "FFF2CC"        # Light yellow
SCHEDULED_COLOR = "C6EFCE"       # Green
PENDING_COLOR = "FFEB9C"         # Yellow


def get_or_create_workbook():
    """
    Opens existing Excel file or creates a new one with headers.
    """
    os.makedirs(OUTPUTS_FOLDER, exist_ok=True)

    if os.path.exists(EXCEL_PATH):
        try:
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
            print(f"   📂 Opened existing Excel file: {EXCEL_PATH}")
            return wb, ws
        except Exception as e:
            print(f"   ⚠️  Could not open existing file: {e}")
            print(f"   📝 Creating new Excel file...")

    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    # --- Write Headers ---
    for col_num, column_name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num, value=column_name)

        # Header styling
        cell.fill = PatternFill(
            start_color=HEADER_BG_COLOR,
            end_color=HEADER_BG_COLOR,
            fill_type="solid"
        )
        cell.font = Font(
            color=HEADER_FONT_COLOR,
            bold=True,
            size=11,
            name="Calibri"
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = get_border()

    # Freeze header row
    ws.freeze_panes = "A2"

    # Set header row height
    ws.row_dimensions[1].height = 35

    print(f"   📝 Created new Excel file: {EXCEL_PATH}")
    return wb, ws


def get_border():
    """
    Returns a thin border style for cells.
    """
    thin = Side(style="thin", color="B8CCE4")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def set_column_widths(ws):
    """
    Sets optimal column widths for readability.
    """
    column_widths = {
        "Date Received":      18,
        "Candidate Name":     22,
        "Email":              28,
        "Phone":              18,
        "Address":            25,
        "Job Role Applied":   25,
        "Application Type":   18,
        "Skills":             45,
        "Resume File":        35,
        "Interview Date":     18,
        "Interview Time":     15,
        "Interview Status":   18,
        "Calendar Event ID":  35,
    }

    for col_num, column_name in enumerate(EXCEL_COLUMNS, 1):
        col_letter = get_column_letter(col_num)
        width = column_widths.get(column_name, 20)
        ws.column_dimensions[col_letter].width = width


def style_data_row(ws, row_num, application_type, interview_status):
    """
    Applies styling to a data row based on application type and status.
    """
    # Choose row background color
    if application_type == "Internship":
        bg_color = INTERNSHIP_COLOR
    elif application_type == "Full-Time":
        bg_color = FULLTIME_COLOR
    else:
        bg_color = ROW_ODD_COLOR if row_num % 2 == 0 else ROW_EVEN_COLOR

    for col_num in range(1, len(EXCEL_COLUMNS) + 1):
        cell = ws.cell(row=row_num, column=col_num)

        # Background color
        cell.fill = PatternFill(
            start_color=bg_color,
            end_color=bg_color,
            fill_type="solid"
        )

        # Font
        cell.font = Font(
            size=10,
            name="Calibri"
        )

        # Alignment
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        # Border
        cell.border = get_border()

    # Special styling for Interview Status column
    status_col = EXCEL_COLUMNS.index("Interview Status") + 1
    status_cell = ws.cell(row=row_num, column=status_col)

    if interview_status == "Scheduled":
        status_cell.fill = PatternFill(
            start_color=SCHEDULED_COLOR,
            end_color=SCHEDULED_COLOR,
            fill_type="solid"
        )
        status_cell.font = Font(
            size=10,
            bold=True,
            color="276221",
            name="Calibri"
        )
    elif interview_status == "Pending":
        status_cell.fill = PatternFill(
            start_color=PENDING_COLOR,
            end_color=PENDING_COLOR,
            fill_type="solid"
        )
        status_cell.font = Font(
            size=10,
            bold=True,
            color="9C5700",
            name="Calibri"
        )

    # Set row height
    ws.row_dimensions[row_num].height = 25


def format_date(date_str):
    """
    Formats date string to a clean readable format.
    """
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(date_str)
        return dt.strftime("%d %b %Y, %I:%M %p")
    except:
        try:
            return datetime.now().strftime("%d %b %Y, %I:%M %p")
        except:
            return str(date_str)


def candidate_already_exists(ws, email):
    """
    Checks if candidate with same email already exists in Excel.
    Prevents duplicate entries.
    """
    email_col = EXCEL_COLUMNS.index("Email") + 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[email_col - 1] == email:
            return True
    return False


def add_candidate_to_excel(candidate_info, interview_info=None):
    """
    Main function — adds candidate data to Excel sheet.
    Takes candidate_info from resume_parser.py
    Takes interview_info from calendar_scheduler.py (optional)
    """
    print(f"\n📊 Saving candidate to Excel...")

    try:
        wb, ws = get_or_create_workbook()

        # Check for duplicate
        if candidate_already_exists(ws, candidate_info.get("email", "")):
            print(f"   ⚠️  Candidate {candidate_info.get('email')} already exists in Excel!")
            print(f"   ⏭️  Skipping duplicate entry")
            return False

        # Get next empty row
        next_row = ws.max_row + 1

        # --- Prepare data ---
        interview_date = ""
        interview_time = ""
        interview_status = "Pending"
        calendar_event_id = ""

        if interview_info:
            interview_date = interview_info.get("date", "")
            interview_time = interview_info.get("time", "")
            interview_status = interview_info.get("status", "Scheduled")
            calendar_event_id = interview_info.get("event_id", "")

        # Resume file — show just filename not full path
        resume_path = candidate_info.get("resume_path", "")
        resume_filename = os.path.basename(resume_path) if resume_path else "No attachment"

        # --- Row data in correct column order ---
        row_data = {
            "Date Received":     format_date(candidate_info.get("date_received", "")),
            "Candidate Name":    candidate_info.get("name", "Unknown"),
            "Email":             candidate_info.get("email", ""),
            "Phone":             candidate_info.get("phone", "Not found"),
            "Address":           candidate_info.get("address", "Not found"),
            "Job Role Applied":  candidate_info.get("job_role", "Not specified"),
            "Application Type":  candidate_info.get("application_type", "Not specified"),
            "Skills":            candidate_info.get("skills", ""),
            "Resume File":       resume_filename,
            "Interview Date":    interview_date,
            "Interview Time":    interview_time,
            "Interview Status":  interview_status,
            "Calendar Event ID": calendar_event_id,
        }

        # --- Write data to row ---
        for col_num, column_name in enumerate(EXCEL_COLUMNS, 1):
            value = row_data.get(column_name, "")
            ws.cell(row=next_row, column=col_num, value=value)

        # --- Apply styling ---
        style_data_row(
            ws,
            next_row,
            candidate_info.get("application_type", ""),
            interview_status
        )

        # --- Set column widths ---
        set_column_widths(ws)

        # --- Save workbook ---
        wb.save(EXCEL_PATH)

        print(f"   ✅ Candidate saved to Excel successfully!")
        print(f"   📍 Row #{next_row - 1} added")
        print(f"   💾 File saved: {EXCEL_PATH}")
        return True

    except Exception as e:
        print(f"   ❌ Error saving to Excel: {e}")
        return False


def update_interview_status(email, new_status, interview_date="", interview_time="", event_id=""):
    """
    Updates interview status for an existing candidate in Excel.
    Useful when rescheduling interviews.
    """
    print(f"\n📊 Updating interview status for: {email}")

    try:
        wb, ws = get_or_create_workbook()

        email_col = EXCEL_COLUMNS.index("Email") + 1
        status_col = EXCEL_COLUMNS.index("Interview Status") + 1
        date_col = EXCEL_COLUMNS.index("Interview Date") + 1
        time_col = EXCEL_COLUMNS.index("Interview Time") + 1
        event_col = EXCEL_COLUMNS.index("Calendar Event ID") + 1

        for row_num in range(2, ws.max_row + 1):
            cell_email = ws.cell(row=row_num, column=email_col).value
            if cell_email == email:
                ws.cell(row=row_num, column=status_col).value = new_status
                if interview_date:
                    ws.cell(row=row_num, column=date_col).value = interview_date
                if interview_time:
                    ws.cell(row=row_num, column=time_col).value = interview_time
                if event_id:
                    ws.cell(row=row_num, column=event_col).value = event_id

                wb.save(EXCEL_PATH)
                print(f"   ✅ Status updated to '{new_status}' for {email}")
                return True

        print(f"   ⚠️  Candidate {email} not found in Excel")
        return False

    except Exception as e:
        print(f"   ❌ Error updating Excel: {e}")
        return False


def print_all_candidates():
    """
    Prints all candidates from Excel to CMD for quick review.
    """
    if not os.path.exists(EXCEL_PATH):
        print("📭 No candidates Excel file found yet.")
        return

    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        total = ws.max_row - 1
        if total <= 0:
            print("📭 No candidates found in Excel.")
            return

        print(f"\n📋 Total candidates in Excel: {total}")
        print("-" * 80)

        name_col = EXCEL_COLUMNS.index("Candidate Name") + 1
        email_col = EXCEL_COLUMNS.index("Email") + 1
        role_col = EXCEL_COLUMNS.index("Job Role Applied") + 1
        type_col = EXCEL_COLUMNS.index("Application Type") + 1
        status_col = EXCEL_COLUMNS.index("Interview Status") + 1
        date_col = EXCEL_COLUMNS.index("Interview Date") + 1

        for row in ws.iter_rows(min_row=2, values_only=True):
            print(f"  👤 {row[name_col-1]}")
            print(f"     📧 {row[email_col-1]}")
            print(f"     💼 {row[role_col-1]} ({row[type_col-1]})")
            print(f"     📅 Interview: {row[date_col-1]} | Status: {row[status_col-1]}")
            print()

    except Exception as e:
        print(f"❌ Error reading Excel: {e}")


def test_excel_writer():
    """
    Test excel writer with dummy data.
    Run this file directly to test.
    """
    print("\n" + "="*50)
    print("  WebEarl HR — Excel Writer Test")
    print("="*50)

    # Dummy candidate data
    test_candidate = {
        "name": "Priyanshu Biswas",
        "email": "p@gmail.com",
        "phone": "+919601599999",
        "address": "Gamdi, Anand, Gujarat",
        "job_role": "AI/ML Developer Intern",
        "application_type": "Internship",
        "skills": "Python, Machine Learning, PyTorch, Flask, Pandas, NumPy",
        "resume_path": "resumes/p_at_gmail_com_20260422_151751.pdf",
        "date_received": "Wed, 22 Apr 2026 15:17:51 +0000",
    }

    # Dummy interview info
    test_interview = {
        "date": "25 Apr 2026",
        "time": "11:00 AM",
        "status": "Scheduled",
        "event_id": "test_event_123",
    }

    success = add_candidate_to_excel(test_candidate, test_interview)

    if success:
        print("\n✅ Test passed! Check outputs/candidates.xlsx")
        print_all_candidates()
    else:
        print("\n❌ Test failed!")


# Run this file directly to test
if __name__ == "__main__":
    test_excel_writer()