# ============================================================
#           WebEarl Technologies — Candidate Response Handler
# ============================================================

import os
import base64
import json
from datetime import datetime, timedelta
from auth import get_credentials
from excel_writer import update_interview_status, EXCEL_PATH
from calendar_scheduler import schedule_interview
from email_sender import send_interview_invitation, build_html_email, build_plain_email, create_email_message, TEMPLATE_PATH
from config import (
    COMPANY_NAME, COMPANY_EMAIL, COMPANY_ADDRESS,
    COMPANY_WEBSITE, OUTPUTS_FOLDER, EXCEL_FILE_NAME
)
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


# ============================================================
#           Response Type Constants
# ============================================================

RESPONSE_CONFIRMED         = "Confirmed"
RESPONSE_RESCHEDULE        = "Reschedule Requested"
RESPONSE_REJECTED          = "Rejected by Candidate"
RESPONSE_QUESTION          = "Question Received"
RESPONSE_NO_REPLY          = "No Response"
RESPONSE_UNRELATED         = "Ignored"

# ============================================================
#           Keywords for Response Detection
# ============================================================

AUTO_REPLY_KEYWORDS = [
    "out of office", "auto reply", "automatic reply",
    "on leave", "on vacation", "away from office",
    "will be back", "currently unavailable",
]


# ============================================================
#           Email Templates for Responses
# ============================================================

CONFIRMATION_ACK_TEMPLATE = os.path.join("templates", "confirmation_ack.html")
RESCHEDULE_TEMPLATE       = os.path.join("templates", "reschedule_invite.html")
REJECTION_ACK_TEMPLATE    = os.path.join("templates", "rejection_ack.html")
REMINDER_TEMPLATE         = os.path.join("templates", "reminder.html")
FOLLOWUP_TEMPLATE         = os.path.join("templates", "followup.html")


# ============================================================
#           Helper — Load HTML Template
# ============================================================

def load_template(template_path, replacements={}):
    """
    Loads an HTML template file and replaces placeholders.
    Falls back to plain text if template not found.
    """
    try:
        with open(template_path, "r", encoding="utf-8") as f:
            html = f.read()
        for key, value in replacements.items():
            html = html.replace(key, str(value))
        return html
    except FileNotFoundError:
        print(f"   ⚠️  Template not found: {template_path}")
        return None
    except Exception as e:
        print(f"   ⚠️  Could not load template: {e}")
        return None


# ============================================================
#           Helper — Send Email
# ============================================================

def send_email(gmail_service, to_email, subject, html_body, plain_body):
    """
    Sends an email via Gmail API.
    """
    try:
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        message = MIMEMultipart("alternative")
        message["To"]      = to_email
        message["From"]    = f"{COMPANY_NAME} <{COMPANY_EMAIL}>"
        message["Subject"] = subject

        message.attach(MIMEText(plain_body, "plain"))
        message.attach(MIMEText(html_body,  "html"))

        raw = base64.urlsafe_b64encode(
            message.as_bytes()
        ).decode("utf-8")

        sent = gmail_service.users().messages().send(
            userId="me",
            body={"raw": raw}
        ).execute()

        print(f"   ✅ Email sent to {to_email} (ID: {sent.get('id', '')})")
        return True

    except Exception as e:
        print(f"   ❌ Could not send email: {e}")
        return False


# ============================================================
#           Helper — Get All Scheduled Candidates from Excel
# ============================================================

def get_scheduled_candidates():
    """
    Reads Excel and returns list of all scheduled candidates.
    """
    candidates = []

    if not os.path.exists(EXCEL_PATH):
        return candidates

    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            candidate = dict(zip(headers, row))
            candidates.append(candidate)

    except Exception as e:
        print(f"   ❌ Could not read Excel: {e}")

    return candidates


# ============================================================
#           Helper — Get Candidate by Email from Excel
# ============================================================

def get_candidate_by_email(email):
    """
    Returns candidate info dict from Excel by email address.
    """
    candidates = get_scheduled_candidates()
    for c in candidates:
        if c.get("Email", "").lower() == email.lower():
            return c
    return None


# ============================================================
#           Step 1 — Detect Response Type
# ============================================================

response_classifier = None

def load_response_classifier():
    global response_classifier
    if response_classifier is None:
        print("   🤖 Loading response classifier model...")
        from transformers import pipeline
        response_classifier = pipeline(
            "zero-shot-classification",
            model="facebook/bart-large-mnli",
            cache_dir=os.path.abspath("models")
        )
        print("   ✅ Response classifier loaded!")
    return response_classifier


def detect_response_type(subject, body):
    text = (subject + " " + body).strip()

    if any(kw in text.lower() for kw in AUTO_REPLY_KEYWORDS):
        return RESPONSE_UNRELATED

    try:
        classifier = load_response_classifier()

        candidate_labels = [
            "candidate is confirming and will attend the interview",
            "candidate wants to reschedule or change the interview date or time",
            "candidate is rejecting cancelling or withdrawing from the interview",
            "candidate is asking a question about the interview",
            "unrelated or spam message",
        ]

        result = classifier(
            text[:1000],
            candidate_labels=candidate_labels,
            multi_label=False
        )

        top_label = result["labels"][0]
        top_score = result["scores"][0]

        print(f"   🤖 AI Detection : {top_label}")
        print(f"   📊 Confidence   : {round(top_score * 100, 1)}%")

        if "confirming" in top_label:
            return RESPONSE_CONFIRMED
        elif "reschedule" in top_label or "change" in top_label:
            return RESPONSE_RESCHEDULE
        elif "rejecting" in top_label or "cancelling" in top_label or "withdrawing" in top_label:
            return RESPONSE_REJECTED
        elif "asking a question" in top_label:
            return RESPONSE_QUESTION
        else:
            return RESPONSE_UNRELATED

    except Exception as e:
        print(f"   ⚠️  AI classification failed: {e}")
        return RESPONSE_UNRELATED

# ============================================================
#           Step 2 — Handle CONFIRMATION
# ============================================================

def handle_confirmation(gmail_service, candidate, sender_email):
    """
    Candidate confirmed — update Excel, send acknowledgement.
    """
    print(f"\n   ✅ Candidate confirmed the interview!")

    # Update Excel status
    update_interview_status(
        sender_email,
        "Confirmed ✅",
    )

    # Load confirmation ack template
    replacements = {
        "{{CANDIDATE_NAME}}" : candidate.get("Candidate Name", "Candidate"),
        "{{JOB_ROLE}}"       : candidate.get("Job Role Applied", "Position"),
        "{{INTERVIEW_DATE}}" : candidate.get("Interview Date", ""),
        "{{INTERVIEW_TIME}}" : candidate.get("Interview Time", ""),
        "{{COMPANY_NAME}}"   : COMPANY_NAME,
        "{{COMPANY_EMAIL}}"  : COMPANY_EMAIL,
        "{{COMPANY_ADDRESS}}": COMPANY_ADDRESS,
        "{{COMPANY_WEBSITE}}": COMPANY_WEBSITE,
    }

    html_body = load_template(CONFIRMATION_ACK_TEMPLATE, replacements)

    if not html_body:
        # Fallback plain HTML
        html_body = f"""
        <p>Dear {candidate.get('Candidate Name', 'Candidate')},</p>
        <p>Thank you for confirming your interview scheduled on
        <strong>{candidate.get('Interview Date', '')}</strong> at
        <strong>{candidate.get('Interview Time', '')}</strong>.</p>
        <p>We will share the meeting link shortly.</p>
        <p>Best Regards,<br><strong>HR Team</strong><br>{COMPANY_NAME}</p>
        """

    plain_body = (
        f"Dear {candidate.get('Candidate Name', 'Candidate')},\n\n"
        f"Thank you for confirming your interview on "
        f"{candidate.get('Interview Date', '')} at "
        f"{candidate.get('Interview Time', '')}.\n\n"
        f"We will share the meeting link shortly.\n\n"
        f"Best Regards,\nHR Team\n{COMPANY_NAME}"
    )

    subject = f"Interview Confirmed — {candidate.get('Job Role Applied', '')} | {COMPANY_NAME}"

    send_email(
        gmail_service,
        sender_email,
        subject,
        html_body,
        plain_body
    )


# ============================================================
#           Step 3 — Handle RESCHEDULE REQUEST
# ============================================================

def handle_reschedule(gmail_service, candidate, sender_email):
    """
    Candidate wants to reschedule — notify HR, delete old event,
    schedule new interview, send new invite.
    """
    print(f"\n   🔄 Candidate requested reschedule!")

    # Update Excel status
    update_interview_status(sender_email, "Reschedule Requested 🔄")

    # Delete old calendar event
    old_event_id = candidate.get("Calendar Event ID", "")
    if old_event_id:
        delete_calendar_event(old_event_id)

    print(f"\n   📅 Please schedule a new interview for this candidate:")
    print(f"   👤 Name   : {candidate.get('Candidate Name', '')}")
    print(f"   💼 Role   : {candidate.get('Job Role Applied', '')}")
    print(f"   📧 Email  : {sender_email}")

    # Build candidate info for rescheduling
    candidate_info = {
        "name"             : candidate.get("Candidate Name", ""),
        "email"            : sender_email,
        "phone"            : candidate.get("Phone", ""),
        "job_role"         : candidate.get("Job Role Applied", ""),
        "application_type" : candidate.get("Application Type", ""),
        "skills"           : candidate.get("Skills", ""),
        "address"          : candidate.get("Address", ""),
    }

    # Schedule new interview
    new_interview = schedule_interview(candidate_info)

    if new_interview:
        # Update Excel with new date/time
        update_interview_status(
            sender_email,
            "Rescheduled ✅",
            interview_date = new_interview.get("date", ""),
            interview_time = new_interview.get("time", ""),
            event_id       = new_interview.get("event_id", ""),
        )

       # --- Send reschedule invite using reschedule_invite.html ---
        print(f"\n📧 Sending reschedule invitation to: {sender_email}")

        replacements = {
            "{{CANDIDATE_NAME}}" : candidate_info["name"],
            "{{JOB_ROLE}}"       : candidate_info["job_role"],
            "{{INTERVIEW_DATE}}" : new_interview.get("date", ""),
            "{{INTERVIEW_TIME}}" : new_interview.get("time", ""),
            "{{COMPANY_NAME}}"   : COMPANY_NAME,
            "{{COMPANY_EMAIL}}"  : COMPANY_EMAIL,
            "{{COMPANY_ADDRESS}}": COMPANY_ADDRESS,
            "{{COMPANY_WEBSITE}}": COMPANY_WEBSITE,
        }

        html_body = load_template(RESCHEDULE_TEMPLATE, replacements)

        if not html_body:
            html_body = f"""
            <p>Dear {candidate_info['name']},</p>
            <p>As requested your interview has been rescheduled to
            <strong>{new_interview.get('date', '')}</strong> at
            <strong>{new_interview.get('time', '')}</strong>.</p>
            <p>Please confirm your availability.</p>
            <p>Best Regards,<br><strong>HR Team</strong><br>{COMPANY_NAME}</p>
            """

        plain_body = (
            f"Dear {candidate_info['name']},\n\n"
            f"As requested your interview has been rescheduled to "
            f"{new_interview.get('date', '')} at "
            f"{new_interview.get('time', '')}.\n\n"
            f"Please confirm your availability.\n\n"
            f"Best Regards,\nHR Team\n{COMPANY_NAME}"
        )

        subject = (
            f"Interview Rescheduled — "
            f"{candidate_info['job_role']} | {COMPANY_NAME}"
        )

        send_email(
            gmail_service,
            sender_email,
            subject,
            html_body,
            plain_body
        )

        print(f"   ✅ Reschedule invitation sent to {sender_email}")
    else:
        print(f"   ⚠️  Could not schedule new interview automatically")


# ============================================================
#           Step 4 — Handle REJECTION
# ============================================================

def handle_rejection(gmail_service, candidate, sender_email):
    """
    Candidate rejected — update Excel, cancel calendar,
    send goodbye email.
    """
    print(f"\n   ❌ Candidate rejected / withdrew application!")

    # Update Excel status
    update_interview_status(sender_email, "Rejected by Candidate ❌")

    # Delete calendar event
    old_event_id = candidate.get("Calendar Event ID", "")
    if old_event_id:
        delete_calendar_event(old_event_id)

    # Send goodbye email
    replacements = {
        "{{CANDIDATE_NAME}}" : candidate.get("Candidate Name", "Candidate"),
        "{{JOB_ROLE}}"       : candidate.get("Job Role Applied", "Position"),
        "{{COMPANY_NAME}}"   : COMPANY_NAME,
        "{{COMPANY_EMAIL}}"  : COMPANY_EMAIL,
        "{{COMPANY_WEBSITE}}": COMPANY_WEBSITE,
    }

    html_body = load_template(REJECTION_ACK_TEMPLATE, replacements)

    if not html_body:
        html_body = f"""
        <p>Dear {candidate.get('Candidate Name', 'Candidate')},</p>
        <p>Thank you for letting us know. We respect your decision and
        wish you all the best in your future endeavours.</p>
        <p>We hope to connect again in the future!</p>
        <p>Best Regards,<br><strong>HR Team</strong><br>{COMPANY_NAME}</p>
        """

    plain_body = (
        f"Dear {candidate.get('Candidate Name', 'Candidate')},\n\n"
        f"Thank you for letting us know. We respect your decision "
        f"and wish you all the best in your future endeavours.\n\n"
        f"We hope to connect again in the future!\n\n"
        f"Best Regards,\nHR Team\n{COMPANY_NAME}"
    )

    subject = f"Thank You for Your Response | {COMPANY_NAME}"

    send_email(
        gmail_service,
        sender_email,
        subject,
        html_body,
        plain_body
    )


# ============================================================
#           Step 5 — Handle QUESTION
# ============================================================

def handle_question(gmail_service, candidate, sender_email, body):
    """
    Candidate asked a question — notify HR and wait for manual reply.
    """
    print(f"\n   ❓ Candidate asked a question!")
    print(f"\n   {'='*50}")
    print(f"   Question from: {candidate.get('Candidate Name', sender_email)}")
    print(f"   {'='*50}")
    print(f"   {body[:500]}")
    print(f"   {'='*50}")

    # Update Excel status
    update_interview_status(sender_email, "Question Received ❓")

    print(f"\n   ⚠️  Please reply manually to: {sender_email}")
    print(f"   💡 You can reply directly from your Gmail")


# ============================================================
#           Step 6 — Delete Calendar Event
# ============================================================

def delete_calendar_event(event_id):
    """
    Deletes a Google Calendar event by event ID.
    """
    try:
        _, _, calendar_service = get_credentials()
        if calendar_service:
            calendar_service.events().delete(
                calendarId="primary",
                eventId=event_id
            ).execute()
            print(f"   🗑️  Calendar event deleted: {event_id}")
    except Exception as e:
        print(f"   ⚠️  Could not delete calendar event: {e}")


# ============================================================
#           Step 7 — Send Reminder to No-Reply Candidates
# ============================================================

def send_reminders(gmail_service):
    """
    Checks Excel for candidates who haven't replied in 2 days.
    Sends reminder or final follow-up email automatically.
    """
    print(f"\n⏰ Checking for candidates with no response...")

    candidates = get_scheduled_candidates()
    today = datetime.today().date()
    reminded = 0

    for candidate in candidates:
        status = candidate.get("Interview Status", "")

        # Only check scheduled candidates
        if status not in ["Scheduled", "No Response ⚠️"]:
            continue

        # Parse interview date
        interview_date_str = candidate.get("Interview Date", "")
        if not interview_date_str:
            continue

        try:
            interview_date = datetime.strptime(
                str(interview_date_str), "%d %b %Y"
            ).date()
        except:
            continue

        days_until = (interview_date - today).days
        candidate_email = candidate.get("Email", "")
        candidate_name  = candidate.get("Candidate Name", "Candidate")
        job_role        = candidate.get("Job Role Applied", "")

        replacements = {
            "{{CANDIDATE_NAME}}" : candidate_name,
            "{{JOB_ROLE}}"       : job_role,
            "{{INTERVIEW_DATE}}" : interview_date_str,
            "{{INTERVIEW_TIME}}" : candidate.get("Interview Time", ""),
            "{{COMPANY_NAME}}"   : COMPANY_NAME,
            "{{COMPANY_EMAIL}}"  : COMPANY_EMAIL,
            "{{COMPANY_WEBSITE}}": COMPANY_WEBSITE,
        }

        # 2 days before — send reminder
        if days_until == 2 and status == "Scheduled":
            print(f"   📨 Sending reminder to {candidate_name}...")

            html_body = load_template(REMINDER_TEMPLATE, replacements)
            if not html_body:
                html_body = f"""
                <p>Dear {candidate_name},</p>
                <p>This is a friendly reminder that your interview for
                <strong>{job_role}</strong> at {COMPANY_NAME} is scheduled on
                <strong>{interview_date_str}</strong> at
                <strong>{candidate.get('Interview Time', '')}</strong>.</p>
                <p>Please confirm your availability by replying to this email.</p>
                <p>Best Regards,<br><strong>HR Team</strong><br>{COMPANY_NAME}</p>
                """

            plain_body = (
                f"Dear {candidate_name},\n\n"
                f"Reminder: Your interview for {job_role} is on "
                f"{interview_date_str} at {candidate.get('Interview Time', '')}.\n\n"
                f"Please confirm your availability.\n\n"
                f"Best Regards,\nHR Team\n{COMPANY_NAME}"
            )

            subject = f"Interview Reminder — {job_role} | {COMPANY_NAME}"
            send_email(gmail_service, candidate_email, subject, html_body, plain_body)
            update_interview_status(candidate_email, "Reminder Sent 📨")
            reminded += 1

        # 1 day before with no response — send final follow up
        elif days_until == 1 and status in ["Scheduled", "Reminder Sent 📨"]:
            print(f"   📨 Sending final follow-up to {candidate_name}...")

            html_body = load_template(FOLLOWUP_TEMPLATE, replacements)
            if not html_body:
                html_body = f"""
                <p>Dear {candidate_name},</p>
                <p>We noticed we have not received a confirmation from you
                regarding your interview tomorrow for
                <strong>{job_role}</strong> at {COMPANY_NAME}.</p>
                <p>Please confirm as soon as possible or let us know if
                you need to reschedule.</p>
                <p>Best Regards,<br><strong>HR Team</strong><br>{COMPANY_NAME}</p>
                """

            plain_body = (
                f"Dear {candidate_name},\n\n"
                f"We have not received your confirmation for the interview "
                f"tomorrow for {job_role}.\n\n"
                f"Please confirm or let us know if you need to reschedule.\n\n"
                f"Best Regards,\nHR Team\n{COMPANY_NAME}"
            )

            subject = f"Final Reminder — Interview Tomorrow | {COMPANY_NAME}"
            send_email(gmail_service, candidate_email, subject, html_body, plain_body)
            update_interview_status(candidate_email, "Final Reminder Sent ⚠️")
            reminded += 1

        # Interview day passed with no response
        elif days_until < 0 and status not in [
            "Confirmed ✅", "Rejected by Candidate ❌",
            "Rescheduled ✅", "No Response ⚠️"
        ]:
            update_interview_status(candidate_email, "No Response ⚠️")
            print(f"   ⚠️  {candidate_name} — marked as No Response")

    if reminded == 0:
        print(f"   ✅ No reminders needed right now")
    else:
        print(f"   ✅ Sent {reminded} reminder(s)")


# ============================================================
#           Step 8 — Fetch Candidate Replies from Gmail
# ============================================================

def get_candidate_replies(gmail_service):
    """
    Fetches unread replies from candidates in Gmail.
    Matches them against scheduled candidates in Excel.
    """
    print(f"\n📬 Scanning Gmail for candidate replies...")

    try:
        # Get scheduled candidate emails from Excel
        candidates = get_scheduled_candidates()
        candidate_emails = [
            c.get("Email", "").lower()
            for c in candidates
            if c.get("Email")
        ]

        if not candidate_emails:
            print(f"   📭 No scheduled candidates found in Excel")
            return []

        # Search unread emails
        result = gmail_service.users().messages().list(
            userId="me",
            q="is:unread",
            maxResults=50
        ).execute()

        messages = result.get("messages", [])
        replies = []

        for msg in messages:
            try:
                full_msg = gmail_service.users().messages().get(
                    userId="me",
                    id=msg["id"],
                    format="full"
                ).execute()

                headers = full_msg["payload"]["headers"]
                header_dict = {h["name"].lower(): h["value"] for h in headers}

                from_header = header_dict.get("from", "")
                sender_email = ""

                if "<" in from_header:
                    sender_email = from_header.split("<")[1].replace(">", "").strip().lower()
                else:
                    sender_email = from_header.strip().lower()

                # Only process if sender is a scheduled candidate
                if sender_email not in candidate_emails:
                    continue

                subject = header_dict.get("subject", "")

                # Extract body
                body = ""
                payload = full_msg["payload"]
                if "parts" in payload:
                    for part in payload["parts"]:
                        if part.get("mimeType") == "text/plain":
                            data = part.get("body", {}).get("data", "")
                            if data:
                                body = base64.urlsafe_b64decode(data).decode("utf-8", errors="ignore")
                                break
                else:
                    data = payload.get("body", {}).get("data", "")
                    if data:
                        body = base64.urlsafe_b64decode(data).decode("utf-8", errors="ignore")

                # Mark as read
                gmail_service.users().messages().modify(
                    userId="me",
                    id=msg["id"],
                    body={"removeLabelIds": ["UNREAD"]}
                ).execute()

                replies.append({
                    "sender_email" : sender_email,
                    "subject"      : subject,
                    "body"         : body.strip(),
                    "message_id"   : msg["id"],
                })

                print(f"   📩 Reply from: {sender_email}")

            except Exception as e:
                print(f"   ⚠️  Error processing message: {e}")
                continue

        print(f"   ✅ Found {len(replies)} candidate reply/replies")
        return replies

    except Exception as e:
        print(f"   ❌ Error fetching replies: {e}")
        return []


# ============================================================
#           Main — Run Response Handler
# ============================================================

def run_response_handler():
    """
    Main function — scans Gmail for candidate replies,
    detects response type and takes appropriate action.
    Also sends reminders to non-responding candidates.
    """
    print("\n" + "="*55)
    print("  WebEarl HR — Candidate Response Handler")
    print("="*55)
    print(f"  Started: {datetime.now().strftime('%d %b %Y, %I:%M %p')}")
    print("="*55)

    # Get services
    _, gmail_service, _ = get_credentials()

    if not gmail_service:
        print("❌ Could not connect to Gmail")
        return

    # -------------------------------------------------------
    # Part 1 — Handle incoming replies
    # -------------------------------------------------------
    replies = get_candidate_replies(gmail_service)

    if replies:
        print(f"\n  📋 Processing {len(replies)} reply/replies...\n")

        for reply in replies:
            sender_email = reply["sender_email"]
            subject      = reply["subject"]
            body         = reply["body"]

            print(f"\n  {'='*55}")
            print(f"  📩 From    : {sender_email}")
            print(f"  📋 Subject : {subject}")
            print(f"  {'='*55}")

            # Get candidate info from Excel
            candidate = get_candidate_by_email(sender_email)

            if not candidate:
                print(f"  ⚠️  Candidate not found in Excel — skipping")
                continue

            # Detect response type
            response_type = detect_response_type(subject, body)
            print(f"  🤖 Detected Response: {response_type}")

            # Handle based on type
            if response_type == RESPONSE_CONFIRMED:
                handle_confirmation(gmail_service, candidate, sender_email)

            elif response_type == RESPONSE_RESCHEDULE:
                handle_reschedule(gmail_service, candidate, sender_email)

            elif response_type == RESPONSE_REJECTED:
                handle_rejection(gmail_service, candidate, sender_email)

            elif response_type == RESPONSE_QUESTION:
                handle_question(gmail_service, candidate, sender_email, body)

            elif response_type == RESPONSE_UNRELATED:
                print(f"  ⏭️  Unrelated reply — skipping")

    else:
        print(f"\n  📭 No new candidate replies found")

    # -------------------------------------------------------
    # Part 2 — Send reminders to non-responding candidates
    # -------------------------------------------------------
    print(f"\n  {'='*55}")
    print(f"  ⏰ Checking reminders...")
    print(f"  {'='*55}")
    send_reminders(gmail_service)

    print(f"\n  {'='*55}")
    print(f"  ✅ Response Handler completed!")
    print(f"  {'='*55}\n")


# ============================================================
#           Entry Point
# ============================================================

if __name__ == "__main__":
    run_response_handler()